"""
utils/excel_parser.py
─────────────────────
Merged-cell-aware parser for the college timetable Excel workbook.

Responsibility (Single Responsibility Principle)
────────────────────────────────────────────────
This module ONLY reads the Excel file and yields raw structured records.
It does NOT touch the database, resolve IDs, or validate against any
lookup table. That separation lives in timetable_import_service.py.

Expected Excel structure (CS-STAFF sheet)
─────────────────────────────────────────
Teacher blocks appear sequentially in the sheet. Each block looks like:

    Row N  :  <Teacher Name>          ← single cell, possibly merged
    Row N+1:  Day/Hour | 1 | 2 | 3 | 4 | 5   ← header (we skip it)
    Row N+2:  I        | c | c | c | c | c    ← Day Order 1, Periods 1-5
    Row N+3:  II       | c | c | c | c | c    ← Day Order 2
    Row N+4:  III      | c | c | c | c | c    ← Day Order 3
    Row N+5:  IV       | c | c | c | c | c    ← Day Order 4
    Row N+6:  V        | c | c | c | c | c    ← Day Order 5
    Row N+7:  VI       | c | c | c | c | c    ← Day Order 6

Each cell `c` may be:
    - Empty/None                → skip
    - "Subject"                 → subject only (no class assignable)
    - "Subject\\nClass"         → subject + class (most common)
    - "Lab Batch\\nClass"       → subject = "Lab Batch", class = ...

Merged cells are transparently resolved: the top-left cell value is used
for every row/column in the merged range.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Generator

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Roman numeral → integer for the Day Order column label
_ROMAN = {"I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6}

# Regex to detect the "Day/Hour" header row that follows a teacher name row
_HEADER_PATTERN = re.compile(r"day\s*/?\s*hour", re.IGNORECASE)

# A teacher block always has exactly this many data rows after the header
DAY_COUNT = 6
PERIOD_COUNT = 5

# Sentinel: the day-order column is column index 0 (first column of the block)
DAY_COL_OFFSET = 0
PERIOD_START_OFFSET = 1  # periods 1..5 are in offsets 1..5


@dataclass
class ParsedSlot:
    """
    Raw slot record produced by the Excel parser before DB resolution.

    Attributes
    ----------
    teacher_name   : raw teacher name string as it appears in the Excel
    day_order      : integer 1-6
    period_number  : integer 1-5
    subject_raw    : first non-empty line from the cell (subject name)
    class_raw      : second non-empty line from the cell, if present
    source_cell    : Excel cell address for traceability (e.g. "C5")
    cell_raw       : full raw cell text for preview and custom parsing
    """

    teacher_name: str
    day_order: int
    period_number: int
    subject_raw: str
    class_raw: str | None
    source_cell: str = field(default="")
    cell_raw: str = field(default="")


# ── Merged-cell resolver ───────────────────────────────────────────────────────

def _build_merge_map(ws) -> dict[tuple[int, int], str | None]:
    """
    Build a (row, col) → cell_value mapping that resolves merged ranges.

    For any cell inside a merged range, returns the value of the top-left
    anchor cell instead of None.
    """
    merge_map: dict[tuple[int, int], str | None] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor_cell = ws.cell(merged_range.min_row, merged_range.min_col)
        anchor_value = anchor_cell.value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merge_map[(row, col)] = anchor_value
    return merge_map


def _get_cell_value(ws, merge_map: dict, row: int, col: int) -> str | None:
    """
    Return the effective string value for a cell, resolving merged ranges.
    Returns None for blank/whitespace-only cells.
    """
    if (row, col) in merge_map:
        raw = merge_map[(row, col)]
    else:
        raw = ws.cell(row, col).value

    if raw is None:
        return None
    text = str(raw).strip()
    return text if text else None


# ── Cell content parser ────────────────────────────────────────────────────────

def _parse_cell_text(text: str) -> tuple[str, str | None]:
    """
    Split a multi-line cell into (subject_name, class_name | None).
    Also handles single-line cells that contain both class and subject (e.g. "III IT A NW" or "II CS D  JAVA").
    """
    lines = [ln.strip() for ln in re.split(r"[\n\r]+", text) if ln.strip()]
    if not lines:
        return ("", None)
        
    if len(lines) >= 2:
        subject = lines[0]
        class_name = lines[1]
        return (subject, class_name)
        
    line = lines[0]
    class_regex = re.compile(
        r"^((?:I{1,3}))" # Roman numeral
        r"\s*[-—]?\s*" # optional divider
        r"("
        r"(?:M\.?\s*Sc\.?|MSC|M\.?\s*COM|MCOM|CS|IT|CT|CD|D)\b" # department
        r"(?:\s*[A-D]\b)?" # section
        r"|"
        r"(?:CS|IT|CT|CD|D)\s*[A-D]\b" # alternative dept + section
        r"|"
        r"CS|IT|CT|CD|D" # fallback
        r")"
        r"(?:\s*[-—]\s*|\s+)?",
        re.IGNORECASE
    )
    match = class_regex.match(line)
    if match:
        class_part = match.group(1) + " " + match.group(2)
        class_part = re.sub(r"\s+", " ", class_part).strip()
        subject_part = line[match.end():].strip()
        subject_part = re.sub(r"^[-—\s\.]+", "", subject_part).strip()
        
        if not subject_part:
            return (class_part, None)
            
        return (subject_part, class_part)
        
    return (line, None)


# ── Teacher block detector ─────────────────────────────────────────────────────

def _is_header_row(ws, merge_map: dict, row: int, start_col: int) -> bool:
    """
    Returns True if the row at `row` looks like the Day/Hour header row,
    i.e. the first cell contains "Day/Hour" (case-insensitive).
    """
    val = _get_cell_value(ws, merge_map, row, start_col)
    return bool(val and _HEADER_PATTERN.search(val))


def _is_roman_row(ws, merge_map: dict, row: int, start_col: int) -> bool:
    """
    Returns True if the first cell of the row contains a Roman numeral
    matching one of the 6 Day Order labels.
    """
    val = _get_cell_value(ws, merge_map, row, start_col)
    if not val:
        return False
    return val.strip().upper() in _ROMAN


# ── Main parser ────────────────────────────────────────────────────────────────

def parse_staff_sheet(
    file_bytes: bytes,
    sheet_name: str = "CS-STAFF",
) -> list[ParsedSlot]:
    """
    Parse the CS-STAFF sheet and return a flat list of ParsedSlot records.

    Parameters
    ----------
    file_bytes : raw bytes of the uploaded .xlsx file
    sheet_name : name of the sheet to parse (default "CS-STAFF")

    Returns
    -------
    list[ParsedSlot]  — one record per non-empty timetable cell

    Raises
    ------
    ValueError  — if the sheet is not found in the workbook
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read Excel file: {exc}") from exc

    # Try exact sheet name first, then case-insensitive search
    ws = None
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        for name in wb.sheetnames:
            if name.strip().upper() == sheet_name.strip().upper():
                ws = wb[name]
                break

    if ws is None:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {available}"
        )

    merge_map = _build_merge_map(ws)
    slots: list[ParsedSlot] = []
    max_row = ws.max_row
    max_col = ws.max_column

    # We scan from column 1; the timetable blocks always start in column 1
    start_col = 1

    row = 1
    while row <= max_row:
        # ── Try to detect a teacher name row ──────────────────────────────────
        # A teacher name row is one where:
        #   (a) the cell is non-empty
        #   (b) the *next* row is the Day/Hour header row
        cell_val = _get_cell_value(ws, merge_map, row, start_col)

        if cell_val and row + 1 <= max_row and _is_header_row(ws, merge_map, row + 1, start_col):
            teacher_name = cell_val.strip()
            logger.debug("Found teacher block: %s at row %d", teacher_name, row)

            # Skip the header row (row + 1)
            data_start_row = row + 2
            data_end_row = data_start_row + DAY_COUNT - 1

            if data_end_row > max_row:
                logger.warning(
                    "Teacher '%s': not enough rows for full timetable block (row %d, max_row %d)",
                    teacher_name, data_start_row, max_row,
                )
                row += 1
                continue

            # ── Parse the 6 × 5 grid ─────────────────────────────────────────
            for day_row_offset in range(DAY_COUNT):
                data_row = data_start_row + day_row_offset
                day_label_val = _get_cell_value(ws, merge_map, data_row, start_col)
                if not day_label_val:
                    continue

                day_order = _ROMAN.get(day_label_val.strip().upper())
                if day_order is None:
                    logger.warning(
                        "Unexpected day label '%s' at row %d — skipping row",
                        day_label_val, data_row,
                    )
                    continue

                for period_idx in range(PERIOD_COUNT):
                    period_number = period_idx + 1
                    cell_col = start_col + PERIOD_START_OFFSET + period_idx
                    cell_text = _get_cell_value(ws, merge_map, data_row, cell_col)

                    if not cell_text:
                        continue  # empty slot — nothing to import

                    subject_raw, class_raw = _parse_cell_text(cell_text)
                    if not subject_raw:
                        continue  # cell had whitespace only

                    cell_addr = f"{get_column_letter(cell_col)}{data_row}"
                    slots.append(
                        ParsedSlot(
                            teacher_name=teacher_name,
                            day_order=day_order,
                            period_number=period_number,
                            subject_raw=subject_raw,
                            class_raw=class_raw,
                            source_cell=cell_addr,
                            cell_raw=cell_text,
                        )
                    )

            # Advance past this entire teacher block
            row = data_end_row + 1
            continue

        row += 1

    logger.info(
        "Excel parsing complete: %d raw slots found across all teacher blocks.",
        len(slots),
    )
    return slots
